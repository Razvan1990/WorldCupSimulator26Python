import os.path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from excel.report_parser import TxtParser
from initialization import constants
from logic.logic import LogicCreator


class ExcelReader(object):

    def __init__(self):
        self.parser = TxtParser()
        self.excel_file = os.path.join(os.getcwd(), "results", "results.xlsx")
        self.text_file_paths = os.path.join(os.getcwd(), "results", "text_files")
        self.logic = LogicCreator()
        self.counter_ranks = 1

    def create_excel_if_missing(self, path):
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            wb.save(path)

    def create_excel_stage_sheets(self, third_place_teams_group):
        list_files = os.listdir(self.text_file_paths)
        print(list_files)
        for file in list_files:
            if file[1:] in constants.GROUPS:
                try:
                    list_matches, list_tables = self.parser.parse_results_stages(
                        os.path.join(self.text_file_paths, file))
                    self.create_excel_if_missing(self.excel_file)
                    report_workbook = openpyxl.load_workbook(self.excel_file)
                    if not file[1:] in report_workbook.sheetnames:
                        report_workbook.create_sheet(file[1:])
                    sheet_active = report_workbook[file[1:]]
                    # remove all pattern fills
                    for row in sheet_active.iter_rows():
                        for cell in row:
                            cell.fill = PatternFill()
                    first_headers = ["A1", "B1"]
                    for i in range(len(first_headers)):
                        sheet_active["{}".format(first_headers[i])].value = constants.EXCEL_HEADERS1[i]
                        # put formating
                        sheet_active["{}".format(first_headers[i])].alignment = Alignment(horizontal="center", )
                        sheet_active["{}".format(first_headers[i])].font = Font(name="Calibri", size=14, bold=True,
                                                                                color='FF000000')
                        sheet_active["{}".format(first_headers[i])].fill = PatternFill(fill_type="solid",
                                                                                       fgColor="D9D9D9")
                    # start writing into scores
                    for j in range(0, len(list_matches[0])):
                        sheet_active["A{}".format(j + 2)].value = list_matches[0][j]
                        sheet_active["A{}".format(j + 2)].alignment = Alignment(horizontal="center")
                        sheet_active["A{}".format(j + 2)].font = Font(name="Calibri", size=11, bold=True,
                                                                      color='FF000000')

                        sheet_active["B{}".format(j + 2)].value = list_matches[1][j]
                        sheet_active["B{}".format(j + 2)].alignment = Alignment(horizontal="center",
                                                                                )
                        sheet_active["B{}".format(j + 2)].font = Font(name="Calibri", size=11, bold=True,
                                                                      color='FF000000')
                    # create table report
                    start_index_row = 1  # row 1
                    stop_index_row = 1  # row 1
                    start_column_index = 6  # F
                    stop_index_column = 11  # K
                    sheet_active.merge_cells(start_row=start_index_row, start_column=start_column_index,
                                             end_row=stop_index_row, end_column=stop_index_column)
                    sheet_active["F1"].value = constants.EXCEL_HEADERS1[2]
                    sheet_active["F1"].alignment = Alignment(horizontal="center")
                    sheet_active["F1"].font = Font(name="Calibri", size=14, bold=True,
                                                   color='FF000000')
                    sheet_active["F1"].fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
                    needed_columns = ["F", "G", "H", "I", "J", "K"]
                    # create headers
                    for i in range(len(needed_columns)):
                        sheet_active["{}2".format(needed_columns[i])].value = constants.EXCEL_HEADERS2[i]
                        # put formating
                        sheet_active["{}2".format(needed_columns[i])].alignment = Alignment(horizontal="center", )
                        sheet_active["{}2".format(needed_columns[i])].font = Font(name="Calibri", size=11, bold=True,
                                                                                  italic=True,
                                                                                  color='FF000000')
                        sheet_active["{}2".format(needed_columns[i])].fill = PatternFill(fill_type="solid",
                                                                                         fgColor="FFF2B2")
                    # add date in standings table
                    row_start = 3
                    counter_list_value = 0
                    print(third_place_teams_group)
                    for column in needed_columns:
                        for i in range(0, len(list_tables[0])):
                            sheet_active["{}{}".format(column, i + row_start)].value = list_tables[counter_list_value][
                                i]
                            sheet_active["{}{}".format(column, i + row_start)].alignment = Alignment(
                                horizontal="center")
                            sheet_active["{}{}".format(column, i + row_start)].font = Font(name="Calibri", size=11,
                                                                                           color='FF000000')
                            if i == 0 or i == 1:
                                sheet_active["{}{}".format(column, i + row_start)].fill = PatternFill(fill_type="solid",
                                                                                                      fgColor="C6EFCE")
                            # check if team on third place is in the lucky losers list - hardcoded
                            if sheet_active["G5"].value in third_place_teams_group:
                                if i == 2:
                                    sheet_active["{}{}".format(column, i + row_start)].fill = PatternFill(
                                        fill_type="solid",
                                        fgColor="FFE699")
                                    # put also on rank
                                    sheet_active["F5"].fill = PatternFill(fill_type="solid", fgColor="FFE699")
                        counter_list_value += 1
                    # align everything
                    self.autofit_columns(sheet_active)
                    report_workbook.save(self.excel_file)

                except:
                    raise Exception("Close excel or check error")

    def create_excel_knockout_stages(self):
        list_files = os.listdir(self.text_file_paths)
        for file in list_files:
            if file[1:] in constants.KNOCKOUTS:
                try:
                    list_matches, qualified = self.parser.parse_results_knockouts(
                        os.path.join(self.text_file_paths, file))
                    report_workbook = openpyxl.load_workbook(self.excel_file)
                    # go to the last sheet
                    if not file[1:] in report_workbook.sheetnames:
                        report_workbook.create_sheet(file[1:])
                    sheet_active = report_workbook[file[1:]]
                    report_workbook.save(self.excel_file)
                    # remove all pattern fills
                    for row in sheet_active.iter_rows():
                        for cell in row:
                            cell.fill = PatternFill()
                    first_headers = ["A1", "B1", "C1"]
                    for i in range(len(first_headers)):
                        sheet_active["{}".format(first_headers[i])].value = constants.EXCEL_HEADERS3[i]
                        # put formating
                        sheet_active["{}".format(first_headers[i])].alignment = Alignment(horizontal="center", )
                        sheet_active["{}".format(first_headers[i])].font = Font(name="Calibri", size=14, bold=True,
                                                                                color='FF000000')
                        sheet_active["{}".format(first_headers[i])].fill = PatternFill(fill_type="solid",
                                                                                       fgColor="D9D9D9")
                    # start writing into scores
                    for j in range(0, len(list_matches[0])):
                        sheet_active["A{}".format(j + 2)].value = list_matches[0][j]
                        sheet_active["A{}".format(j + 2)].alignment = Alignment(horizontal="center")
                        sheet_active["A{}".format(j + 2)].font = Font(name="Calibri", size=11, bold=True,
                                                                      color='FF000000')

                        sheet_active["B{}".format(j + 2)].value = list_matches[1][j]
                        sheet_active["B{}".format(j + 2)].alignment = Alignment(horizontal="center",
                                                                                )
                        sheet_active["B{}".format(j + 2)].font = Font(name="Calibri", size=11, bold=True,
                                                                      color='FF000000')

                        sheet_active["C{}".format(j + 2)].value = list_matches[2][j]
                        sheet_active["C{}".format(j + 2)].alignment = Alignment(horizontal="center",
                                                                                )
                        sheet_active["C{}".format(j + 2)].font = Font(name="Calibri", size=11, bold=True, italic=True,
                                                                      color='FF000000')
                    # write the standings
                    if file[1:] != constants.KNOCKOUTS[len(constants.KNOCKOUTS) - 1] and file[1:] != \
                            constants.KNOCKOUTS[len(constants.KNOCKOUTS) - 2] and file[1:] != constants.KNOCKOUTS[
                        len(constants.KNOCKOUTS) - 3]:
                        sheet_active["F1"].value = constants.EXCEL_HEADERS3[3]
                    elif file[1:] == constants.KNOCKOUTS[len(constants.KNOCKOUTS) - 2]:
                        sheet_active["F1"].value = constants.SPECIAL_EXCEL_HEADERS[0]
                    elif file[1:] == constants.KNOCKOUTS[len(constants.KNOCKOUTS) - 1]:
                        sheet_active["F1"].value = constants.SPECIAL_EXCEL_HEADERS[1]
                    elif file[1:] == constants.KNOCKOUTS[len(constants.KNOCKOUTS) - 3]:
                        sheet_active["F1"].value = constants.SPECIAL_EXCEL_HEADERS[2]
                        sheet_active["G1"].value = constants.SPECIAL_EXCEL_HEADERS[3]
                        sheet_active["G1"].alignment = Alignment(horizontal="center")
                        sheet_active["G1"].font = Font(name="Calibri", size=11, bold=True,
                                                       color='FF000000')
                        sheet_active["G1"].fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
                    sheet_active["F1"].alignment = Alignment(horizontal="center")
                    sheet_active["F1"].font = Font(name="Calibri", size=11, bold=True,
                                                   color='FF000000')
                    sheet_active["F1"].fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
                    for j in range(0, len(qualified)):
                        # add qualified in small final and big final
                        if file[1:] == constants.KNOCKOUTS[len(constants.KNOCKOUTS) - 3]:
                            if j in [0, 1]:
                                sheet_active["G{}".format(j + 2)].value = qualified[j]
                                sheet_active["G{}".format(j + 2)].alignment = Alignment(horizontal="center")
                                sheet_active["G{}".format(j + 2)].font = Font(name="Calibri", size=11, bold=True,
                                                                              italic=True,
                                                                              color='FF000000')
                                sheet_active["G{}".format(j + 2)].fill = PatternFill(fill_type="solid",
                                                                                     fgColor="C6EFCE")
                            else:
                                # we put j as we are already at second index
                                sheet_active["F{}".format(j)].value = qualified[j]
                                sheet_active["F{}".format(j)].alignment = Alignment(horizontal="center")
                                sheet_active["F{}".format(j)].font = Font(name="Calibri", size=11, bold=True,
                                                                          italic=True,
                                                                          color='FF000000')
                                sheet_active["F{}".format(j)].fill = PatternFill(fill_type="solid",
                                                                                 fgColor="FF9999")
                        else:
                            sheet_active["F{}".format(j + 2)].value = qualified[j]
                            sheet_active["F{}".format(j + 2)].alignment = Alignment(horizontal="center")
                            sheet_active["F{}".format(j + 2)].font = Font(name="Calibri", size=11, bold=True,
                                                                          italic=True,
                                                                          color='FF000000')
                            sheet_active["F{}".format(j + 2)].fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
                    # align everything
                    self.autofit_columns(sheet_active)
                    report_workbook.save(self.excel_file)
                except:
                    print("Close excel or check error")

    def create_sheet_final_rankings_sheet(self):
        try:
            file = os.listdir(self.text_file_paths)[len(os.listdir(self.text_file_paths)) - 2] #need to check here if we had more files
            rankings = self.parser.parse_final_stats_file(os.path.join(self.text_file_paths, file))
            report_workbook = openpyxl.load_workbook(self.excel_file)
            # go to the last sheet
            if not file[1:] in report_workbook.sheetnames:
                report_workbook.create_sheet(file[1:])
            sheet_active = report_workbook[file[1:]]
            report_workbook.save(self.excel_file)
            # remove all pattern fills
            for row in sheet_active.iter_rows():
                for cell in row:
                    cell.fill = PatternFill()
            needed_colummns = ["A", "B", "C", "D"]
            for i in range(len(needed_colummns)):
                sheet_active["{}1".format(needed_colummns[i])].value = constants.FINAL_RANKING_EXCEL_HEADERS[i]
                # put formating
                sheet_active["{}1".format(needed_colummns[i])].alignment = Alignment(horizontal="center", )
                sheet_active["{}1".format(needed_colummns[i])].font = Font(name="Calibri", size=14, bold=True,
                                                                           color='FF000000')
                sheet_active["{}1".format(needed_colummns[i])].fill = PatternFill(fill_type="solid",
                                                                                  fgColor="D9D9D9")
            # populate list
            counter_lists = 0
            while counter_lists < len(rankings):
                for i in range(len(needed_colummns)):
                    sheet_active["{}{}".format(needed_colummns[i], counter_lists + 2)].value = rankings[counter_lists][
                        i]
                    sheet_active["{}{}".format(needed_colummns[i], counter_lists + 2)].alignment = Alignment(
                        horizontal="center", )
                    sheet_active["{}{}".format(needed_colummns[i], counter_lists + 2)].font = Font(name="Calibri",
                                                                                                   size=11, bold=True, )
                counter_lists += 1
            # align everything
            self.autofit_columns(sheet_active)
            report_workbook.save(self.excel_file)

        except:
            raise Exception("Close excel or check error")

    def create_sheet_strikers_list(self):
        try:
            file = os.listdir(self.text_file_paths)[len(os.listdir(self.text_file_paths)) - 1] #need to check here if we had more files
            player_goals = self.parser.parse_goals_players_file(os.path.join(self.text_file_paths, file))
            report_workbook = openpyxl.load_workbook(self.excel_file)
            # go to the last sheet
            if not file[1:] in report_workbook.sheetnames:
                report_workbook.create_sheet(file[1:])
            sheet_active = report_workbook[file[1:]]
            report_workbook.save(self.excel_file)
            # remove all pattern fills
            for row in sheet_active.iter_rows():
                for cell in row:
                    cell.fill = PatternFill()

            # create headers
            needed_colummns = ["A", "B", "C", "D"]
            for i in range(len(needed_colummns)):
                sheet_active["{}1".format(needed_colummns[i])].value = constants.PLAYER_HEADERS[i]
                # put formating
                sheet_active["{}1".format(needed_colummns[i])].alignment = Alignment(horizontal="center", )
                sheet_active["{}1".format(needed_colummns[i])].font = Font(name="Calibri", size=14, bold=True,
                                                                           color='FF000000')
                sheet_active["{}1".format(needed_colummns[i])].fill = PatternFill(fill_type="solid",
                                                                                  fgColor="D9D9D9")
            # start populating the list
            counter_lists = 0
            while counter_lists < len(player_goals):
                for i in range(len(needed_colummns)):
                    if counter_lists in [0, 1, 2]:
                        sheet_active["{}{}".format(needed_colummns[i], counter_lists + 2)].fill = PatternFill(
                            fill_type="solid",fgColor="C6EFCE")
                    if needed_colummns[i] =="A":
                        sheet_active["{}{}".format(needed_colummns[i], counter_lists + 2)].value = str(self.counter_ranks)
                        self.counter_ranks += 1
                    else:
                        sheet_active["{}{}".format(needed_colummns[i], counter_lists + 2)].value = player_goals[counter_lists][i-1] #we will not have index out of range because here starts from index 2
                    # put formating
                    sheet_active["{}1".format(needed_colummns[i])].alignment = Alignment(horizontal="center", )
                    sheet_active["{}1".format(needed_colummns[i])].font = Font(name="Calibri", size=14, bold=True,
                                                                               color='FF000000')
                counter_lists += 1
                # align everything
            self.autofit_columns(sheet_active)
            report_workbook.save(self.excel_file)

        except:
            raise Exception("Close excel or check error")


    '''
    COPILOT CODE
    '''

    def autofit_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # numeric index
            column_letter = get_column_letter(column)

            for cell in col:
                try:
                    cell_value = str(cell.value)
                    if cell_value:
                        max_length = max(max_length + 1, len(cell_value))
                except:
                    pass

            # Add a little padding
            adjusted_width = max_length
            ws.column_dimensions[column_letter].width = adjusted_width
